import pandas as pd
import json
from openpyxl.utils import get_column_letter
from datetime import datetime

def serialize_datetime(value):
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    return value

from openpyxl import load_workbook

class ExcelParser:
    """
    Class for parsing Excel files with two tables: summary and detailed.

    Attributes:
    filepath (str): Path to Excel file.
    wb (Workbook): Opened Excel workbook via openpyxl (to access formulas).

    Methods:
    parse(): Returns a list of documents extracted from all tables.
    save_as_jsonl(docs, output_path): Saves a list of documents in JSONL format.
    """

    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = load_workbook(filepath, data_only=False)


    def parse(self):
        xls = pd.ExcelFile(self.filepath)
        documents = []

        sheet1_name = xls.sheet_names[0]
        df_sheet1 = pd.read_excel(xls, sheet_name=sheet1_name, header=None)

        data1 = self._parse_summary_sheet(df_sheet1, sheet1_name, start_row=3, group_row=2)
        documents.extend(data1)

        data2 = self._parse_summary_sheet(df_sheet1, sheet1_name, start_row=19, group_row=18)
        documents.extend(data2)

        data3 = self._parse_details_sheet(self.filepath, sheet_name=xls.sheet_names[1])
        documents.extend(data3)

        return documents

    def _parse_summary_sheet(self, df: pd.DataFrame, sheet_name: str, start_row: int, group_row: int) -> list[dict]:
        """
        Parses a pivot table row by row, associating metrics with groups and values.

        Args:
        df(pd.DataFrame): DataFrame with sheet contents.
        sheet_name(str): Sheet name.
        start_row(int): Start row of data (0-indexed).
        group_row(int): String containing group names (0-indexed).

        Returns:
        list[dict]: List of documents row by row.
        """

        jsonl_data = []
        start_col = 2
        ws = self.wb[sheet_name]

        for row_index in range(start_row, df.shape[0]):
            metric = df.iat[row_index, 1]
            if pd.isna(metric):
                continue
            metric = str(metric).strip()

            text_parts = [f"{sheet_name} | {metric}"]
            meta_entries = []

            for col_index in range(start_col, df.shape[1]):
                group = df.iat[group_row, col_index]
                value = df.iat[row_index, col_index]

                if pd.isna(group) or pd.isna(value):
                    continue

                group = str(group).strip()
                col_letter = get_column_letter(col_index + 1)
                cell = ws[f"{col_letter}{row_index + 1}"]
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None

                try:
                    value_float = float(value)
                    value_str = str(round(value_float, 6))
                except (ValueError, TypeError):
                    value_str = str(value)

                formula_part = f" (формула: {formula})" if formula else ""
                text_parts.append(f"для {group} = {value_str}{formula_part}")

                meta_entries.append({
                    "group": group,
                    "col_index": col_letter,
                    "value": serialize_datetime(value),
                    "formula": formula
                })

            if len(meta_entries) == 0:
                continue

            text = " ".join(text_parts)
            metadata = {
                "type": "by_row",
                "sheet": sheet_name,
                "metric": metric,
                "row_index": row_index,
                "entries": meta_entries
            }

            jsonl_data.append({
                "text": text,
                "metadata": metadata
            })

        return jsonl_data

    def _parse_details_sheet(self, file_path, sheet_name="2", header_row_1_idx=3, header_row_2_idx=4, data_start_idx=6):
        """
        Parses a table with details using a double header.

        Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet.
        header_row_1_idx (int): Index of the first header row.
        header_row_2_idx (int): Index of the second header row.
        data_start_idx (int): Index of the row where the data starts.

        Returns:
        list[dict]: List of documents by row.
        """

        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        ws = self.wb[sheet_name]

        header_row_1 = df.iloc[header_row_1_idx].ffill()
        header_row_2 = df.iloc[header_row_2_idx].fillna("")

        jsonl_data = []

        for row_index in range(data_start_idx, len(df)):
            row = df.iloc[row_index]
            entries = []
            text_parts = []

            for col_index, (col1, col2, value) in enumerate(zip(header_row_1, header_row_2, row)):
                col_letter = get_column_letter(col_index + 1)
                group_name = f"{col1} {col2}".strip()
                group_name = " ".join(group_name.split())

                if group_name == "" or pd.isna(value):
                    continue

                cell = ws[f"{col_letter}{row_index + 1}"]
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None

                value_serialized = serialize_datetime(value)
                formula_part = f" (формула: {formula})" if formula else ""

                text_parts.append(f"{group_name} = {value_serialized}{formula_part}")
                entries.append({
                    "group": group_name,
                    "col_index": col_letter,
                    "value": value_serialized,
                    "formula": formula
                })

            if entries:
                text = "Таблица 2 | " + " | ".join(text_parts)
                json_obj = {
                    "text": text,
                    "metadata": {
                        "type": "by_row",
                        "sheet": sheet_name,
                        "row_index": row_index,
                        "entries": entries
                    }
                }
                jsonl_data.append(json_obj)

        return jsonl_data

    def save_as_jsonl(self, docs, output_path):
        with open(output_path, "w", encoding="utf-8") as f:
            for doc in docs:
                f.write(json.dumps(doc, ensure_ascii=False, default=str) + "\n")

